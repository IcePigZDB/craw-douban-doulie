import re
import openpyxl
import requests
from requests import RequestException
from bs4 import BeautifulSoup
import lxml
import time
import random

def get_one_page(url):
    '''
    Get the html of a page by requests module
    @url: page url
    @return: html / None
    '''
    try:
        head = ['Mozilla/5.0', 'Chrome/78.0.3904.97', 'Safari/537.36']
        headers = {
            'user-agent':head[random.randint(0, 2)]
        }
        response = requests.get(url, headers=headers, proxies={'http':'171.15.65.195:9999'}) # 这里的代理，可以设置也可以不加，如果失效，不加或者替换其他的即可
        if response.status_code == 200:
            return response.text
        return None
    except RequestException:
        return None


def get_atr_by_re(pattern_text, search_str):
    '''
    Use re pattern to extra attribute
    @pattern_text: re pattern
    @html: page's html text
    @return: attribute 
    '''
    pattern = re.compile(pattern_text, re.S)
    res = re.findall(pattern, search_str)
    if len(res) > 0 :
        res = re.findall(pattern, search_str)[0]
        res = res.replace(" ", "")
        res = res.replace("\n", "")
        return res
    else:
        return 'NULL'
def write_lable_excel(label,file):
    '''
    Write excel label to excel file
    @label:  lable list in order
    @file: excel filename
    '''
    wb = openpyxl.load_workbook(file)
    ws = wb.worksheets[0]
    # excel start from 1!
    # cell (x,y) location 
    for i,val in enumerate(label):
        ws.cell(1,i+1).value = val
    wb.save(file)

def write_bookinfo_excel(book_info, file):
    '''
    Write book info into excel file
    @book_info: a dict
    @file: excel filename
    @return: the num of successful item
    '''
    wb = openpyxl.load_workbook(file)
    # first sheet,default 3 sheets
    ws = wb.worksheets[0]
    sheet_row = ws.max_row
    sheet_col = ws.max_column
    i = sheet_row
    j = 1
    for key in book_info:
        # append write
        ws.cell(i+1, j).value = book_info[key]
        j += 1
    done = ws.max_row - sheet_row
    wb.save(file)
    return done

def get_info_dict_list(html):
    '''
    Craw wanted infomation for a page
    @html: html used for craw
    @return: a list of info dict,a dict per item
    '''
    info_dict_list = []
    # none return empty
    if html ==None:
        return info_dict_list
    # parser html by lxml
    soup = BeautifulSoup(html, 'lxml')
    # class = bd 
    selector = '.bd'
    bd_class_list = soup.select(selector)
    # if use .bd,there are 2 class = "bd" & 25 class = "bd doulist-subject" per page 
    # so we start from 2, to len(res)
    for i in range(2,len(bd_class_list)):
        info_dict = {}
        # title
        search_str = bd_class_list[i].select(".title > a ")
        pattern_text = '>(.*?)</a>'
        # pattern = re.compile(pattern_text, re.S)
        # title = re.findall(pattern,str(title))[0]
        # title = title.replace(" ", "")
        # title = title.replace("\n", "")
        title  = get_atr_by_re(pattern_text,str(search_str))
        info_dict['title'] = title

        # writer
        search_str = bd_class_list[i].select(".abstract")
        pattern_text = '作者:(.*?)<br/>'
        writer = get_atr_by_re(pattern_text,str(search_str))
        info_dict['writer'] = writer
        # publisher
        pattern_text = '出版社:(.*?)<br/>'
        publisher = get_atr_by_re(pattern_text,str(search_str))
        info_dict['publisher'] = publisher
        # publish year
        pattern_text = '出版年:(.*?)</div>'
        year = get_atr_by_re(pattern_text,str(search_str))
        info_dict['year'] = year

        # rating
        search_str = bd_class_list[i].select(".rating")
        pattern_text = 'rating_nums">(.*?)</span>'
        rating = get_atr_by_re(pattern_text,str(search_str))
        info_dict['rating'] = rating

        info_dict_list.append(info_dict)
        # link 
        search_str = bd_class_list[i].select(".post")
        pattern_text = 'href="(.*?)"'
        link = get_atr_by_re(pattern_text,str(search_str))
        info_dict['link'] = link
        # img source
        pattern_text = 'img src="(.*?)"'
        img = get_atr_by_re(pattern_text,str(search_str))
        info_dict['img'] = img 

    return info_dict_list

# config 
# page number need to be craw
pagenumber = 4
filename = 'dscl_info.xlsx'
# url config below , code may need change!

if __name__ == '__main__':
    total = 0
    for page_index in range(0, pagenumber): 
        # url
        url = 'https://www.douban.com/doulist/135073077/?start='+str(page_index*25)+'&sort=time&playable=0&sub_type='
        one_loop_done = 0
        # only get html page once
        html = get_one_page(url)
        # every dict contain an item infomation
        info_dict_list = get_info_dict_list(html)
        # write to output excel file
        # title 
        write_lable_excel(["title","writer","publisher","publish date","score","link","img source"],filename)
        # book info
        for i in range(len(info_dict_list)):
            row = write_bookinfo_excel(info_dict_list[i], filename)
            one_loop_done += row
        total += one_loop_done
        print(one_loop_done, 'done')

    print('Total', total, 'done')